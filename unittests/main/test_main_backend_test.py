# Copyright (c) ONNX Project Contributors
#
# SPDX-License-Identifier: Apache-2.0
"""Tests for ``python -m onnx_light backend``."""

from __future__ import annotations

import io
import json
import os
import subprocess
import sys
import tempfile
import unittest
from contextlib import redirect_stderr, redirect_stdout
from unittest import mock

from onnx_light.__main__ import _build_parser, _run_backend_test_timing, main
from onnx_light.ext_test_case import ExtTestCase, import_or_skip

import_or_skip("onnx_light.onnx_py._onnxpybackend", "backend_test")
import_or_skip("onnx_light.onnx_py._onnxpykernels", "runtime")


class TestMainBackendTest(ExtTestCase):
    def test_parser_defaults(self):
        args = _build_parser().parse_args(["backend"])
        self.assertEqual(args.regex, "")
        self.assertEqual(args.mode, "test")
        self.assertEqual(args.repeat, 10 * (os.cpu_count() or 1))
        self.assertEqual(args.warmup, 2 * (os.cpu_count() or 1))
        self.assertEqual(args.max_repeat_time, 1.0)
        self.assertEqual(args.timeout, 2.0)
        self.assertIsNone(args.parameter)
        self.assertIsNone(args.criterion)
        self.assertIsNone(args.kernel)
        self.assertIsNone(args.dtype)
        self.assertIsNone(args.impl)
        self.assertFalse(args.include_big)
        self.assertFalse(args.json)
        self.assertIsNone(args.output)
        self.assertIsNone(args.save_models)

    def test_parser_rejects_invalid_counts(self):
        parser = _build_parser()
        with self.assertRaises(SystemExit):
            parser.parse_args(["backend", "--repeat", "0"])
        with self.assertRaises(SystemExit):
            parser.parse_args(["backend", "--warmup", "-1"])
        with self.assertRaises(SystemExit):
            parser.parse_args(["backend", "--max-repeat-time", "0"])
        with self.assertRaises(SystemExit):
            parser.parse_args(["backend", "--timeout", "0"])
        for value in ("nan", "inf"):
            with self.subTest(timeout=value), self.assertRaises(SystemExit):
                parser.parse_args(["backend", "--timeout", value])

    def test_times_one_correctness_case(self):
        report = _run_backend_test_timing(name_regex=r"^test_cc_abs$", repeat=1)
        self.assertEqual(report["mode"], "test")
        self.assertEqual(report["selected"], 1)
        self.assertEqual(report["cases"][0]["name"], "test_cc_abs")
        self.assertEqual(report["cases"][0]["data_sets"], 1)
        self.assertEqual(report["cases"][0]["input_shapes"], [{"x": [2, 3]}])
        self.assertEqual(report["cases"][0]["status"], "completed")
        self.assertFalse(report["cases"][0]["timed_out"])
        self.assertEqual(len(report["cases"][0]["iteration_seconds"]), 1)
        self.assertGreaterEqual(report["cases"][0]["run_seconds"], 0)

    def test_times_one_benchmark_case(self):
        report = _run_backend_test_timing(
            name_regex=r"^test_cc_abs_benchmark$", mode="benchmark", repeat=1
        )
        self.assertEqual(report["mode"], "benchmark")
        self.assertEqual(report["selected"], 1)
        self.assertEqual(report["cases"][0]["name"], "test_cc_abs_benchmark")
        self.assertGreater(report["cases"][0]["run_seconds"], 0)

    def test_marks_case_exceeding_timeout(self):
        report = _run_backend_test_timing(
            name_regex=r"^test_cc_abs$", repeat=1, timeout_seconds=0.000001
        )
        self.assertEqual(report["selected"], 1)
        self.assertEqual(report["timed_out"], 1)
        self.assertEqual(report["cases"][0]["status"], "timeout")
        self.assertTrue(report["cases"][0]["timed_out"])
        self.assertIsNone(report["cases"][0]["run_seconds"])

    def test_limits_each_repeat_phase_by_duration(self):
        report = _run_backend_test_timing(
            name_regex=r"^test_cc_abs$", repeat=10, warmup=10, max_repeat_time=0.000001
        )
        self.assertEqual(len(report["cases"][0]["iteration_seconds"]), 1)
        self.assertGreater(report["cases"][0]["warmup_seconds"], 0)

    def test_json_output(self):
        expected = {
            "name_regex": "abs",
            "mode": "benchmark",
            "include_big": False,
            "repeat": 2,
            "warmup": 1,
            "max_repeat_time": 0.5,
            "timeout_seconds": 3.5,
            "selected": 0,
            "timed_out": 0,
            "collection_seconds": 0.1,
            "cases": [],
            "total_seconds": 0.2,
        }
        buffer = io.StringIO()
        with (
            mock.patch(
                "onnx_light.__main__._run_backend_test_timing", return_value=expected
            ) as run_timing,
            redirect_stdout(buffer),
        ):
            main(
                [
                    "backend",
                    "--regex",
                    "abs",
                    "--mode",
                    "benchmark",
                    "--repeat",
                    "2",
                    "--warmup",
                    "1",
                    "--max-repeat-time",
                    "0.5",
                    "--timeout",
                    "3.5",
                    "--json",
                ]
            )
        self.assertEqual(json.loads(buffer.getvalue()), expected)
        run_timing.assert_called_once_with(
            name_regex="abs",
            mode="benchmark",
            include_big=False,
            generate_benchmark_expected_outputs=False,
            repeat=2,
            warmup=1,
            max_repeat_time=0.5,
            timeout_seconds=3.5,
            tuning_comparison=None,
            save_models=None,
            progress=None,
        )

    def test_rejects_invalid_counts(self):
        with self.assertRaisesRegex(ValueError, "repeat must be positive"):
            _run_backend_test_timing(repeat=0)
        with self.assertRaisesRegex(ValueError, "warmup must be non-negative"):
            _run_backend_test_timing(warmup=-1)
        with self.assertRaisesRegex(ValueError, "max_repeat_time must be positive"):
            _run_backend_test_timing(max_repeat_time=0)
        with self.assertRaisesRegex(ValueError, "timeout_seconds must be positive"):
            _run_backend_test_timing(timeout_seconds=0)
        for value in (float("nan"), float("inf")):
            with (
                self.subTest(timeout=value),
                self.assertRaisesRegex(ValueError, "timeout_seconds must be positive and finite"),
            ):
                _run_backend_test_timing(timeout_seconds=value)

    def test_rejects_invalid_regex(self):
        with self.assertRaises(ValueError):
            _run_backend_test_timing(name_regex="(")

    def test_python_module_command(self):
        process = subprocess.run(
            [sys.executable, "-m", "onnx_light", "backend", "--regex", "^test_cc_abs$", "--json"],
            check=True,
            capture_output=True,
            text=True,
        )
        report = json.loads(process.stdout)
        self.assertEqual(report["selected"], 1)
        self.assertEqual(report["cases"][0]["name"], "test_cc_abs")

    def test_writes_csv_or_xlsx_based_on_extension(self):
        expected = {
            "name_regex": "not",
            "mode": "test",
            "include_big": False,
            "repeat": 2,
            "warmup": 1,
            "max_repeat_time": 0.5,
            "timeout_seconds": 2.0,
            "selected": 1,
            "timed_out": 0,
            "collection_seconds": 0.1,
            "cases": [
                {
                    "name": "test_cc_not",
                    "kind": "node",
                    "tag": "Not",
                    "status": "completed",
                    "timed_out": False,
                    "error": None,
                    "parameter_name": None,
                    "parameter_value": None,
                    "baseline_value": None,
                    "speedup": None,
                    "data_sets": 1,
                    "input_shapes": [{"x": [2, 2]}],
                    "model_path": None,
                    "materialization_seconds": 0.2,
                    "setup_seconds": 0.3,
                    "warmup_seconds": 0.4,
                    "iteration_seconds": [0.5, 0.6],
                    "run_seconds": 1.1,
                    "mean_seconds": 0.55,
                    "min_seconds": 0.5,
                    "max_seconds": 0.6,
                }
            ],
            "total_seconds": 2.1,
            "tuning_comparison": None,
            "save_models": None,
        }
        with tempfile.TemporaryDirectory() as temporary:
            for extension in (".csv", ".xlsx"):
                with self.subTest(extension=extension):
                    output = os.path.join(temporary, f"not{extension}")
                    with (
                        mock.patch(
                            "onnx_light.__main__._run_backend_test_timing", return_value=expected
                        ),
                        redirect_stdout(io.StringIO()),
                    ):
                        main(["backend", "--regex", "not", "--output", output])
                    self.assertTrue(os.path.exists(output))

            with open(os.path.join(temporary, "not.csv"), encoding="utf-8") as csv_file:
                csv_text = csv_file.read()
            self.assertIn(
                "name,kind,tag,status,timed_out,error,parameter_name,"
                "parameter_value,baseline_value,speedup,data_sets,input_shapes",
                csv_text,
            )
            self.assertIn(
                'test_cc_not,node,Not,completed,False,,,,,,1,"[{""x"": [2, 2]}]"', csv_text
            )

            from openpyxl import load_workbook

            workbook = load_workbook(os.path.join(temporary, "not.xlsx"), read_only=True)
            try:
                summary = workbook["summary"]
                summary_values = dict(summary.iter_rows(min_row=2, values_only=True))
                self.assertEqual(summary_values["repeat"], 2)
                self.assertEqual(summary_values["warmup"], 1)
                self.assertEqual(summary_values["max_repeat_time"], 0.5)
                self.assertEqual(summary_values["timeout_seconds"], 2)
                self.assertEqual(summary_values["timed_out"], 0)
                self.assertIn("cpu.architecture", summary_values)
                self.assertIn("cpu.logical_cores", summary_values)
                worksheet = workbook["backend"]
                self.assertEqual(worksheet["A1"].value, "name")
                self.assertEqual(worksheet["A2"].value, "test_cc_not")
            finally:
                workbook.close()

    def test_parameter_comparison_selects_one_tuning_schema(self):
        tunable = {
            "library": "onnx_light",
            "kernel": "Not",
            "implementation": "portable",
            "element_type": 9,
            "device": -1,
            "device_name": "CPU",
            "tuning_abi": 1,
            "calibratable": True,
            "defaults": {"parallel.minimum_elements": 32768},
            "active_values": {"parallel.minimum_elements": 32768},
            "parameter_names": ["parallel.minimum_elements"],
        }
        expected = {"cases": [], "tuning_comparison": None}
        with (
            mock.patch(
                "onnx_light.kernel_tuning.kernel_tuning_parameters",
                return_value={"kernels": [tunable]},
            ),
            mock.patch(
                "onnx_light.__main__._run_backend_test_timing", return_value=expected
            ) as run_timing,
            redirect_stdout(io.StringIO()),
        ):
            main(
                [
                    "backend",
                    "--regex",
                    "not",
                    "--kernel",
                    "Not",
                    "--dtype",
                    "BOOL",
                    "--impl",
                    "portable",
                    "--parameter",
                    "parallel.minimum_elements=default,16384",
                    "--criterion",
                    "average",
                    "--json",
                ]
            )
        comparison = run_timing.call_args.kwargs["tuning_comparison"]
        self.assertEqual(
            comparison["parameter_sets"],
            [{"parallel.minimum_elements": 32768}, {"parallel.minimum_elements": 16384}],
        )
        self.assertEqual(comparison["criterion"], "average")
        self.assertEqual(comparison["tunable"], tunable)

    def test_parameter_comparison_builds_cartesian_product(self):
        tunable = {
            "library": "onnx_light",
            "kernel": "Gemm",
            "implementation": "portable",
            "element_type": 1,
            "device": -1,
            "device_name": "CPU",
            "tuning_abi": 1,
            "calibratable": True,
            "defaults": {"algorithm.tile": 8, "parallel.minimum_tasks": 2},
            "active_values": {"algorithm.tile": 8, "parallel.minimum_tasks": 2},
            "parameter_names": ["algorithm.tile", "parallel.minimum_tasks"],
        }
        expected = {"cases": [], "tuning_comparison": None}
        progress_output = io.StringIO()

        def run_timing(**kwargs):
            kwargs["progress"](1, 4)
            kwargs["progress"](4, 4)
            return expected

        with (
            mock.patch(
                "onnx_light.kernel_tuning.kernel_tuning_parameters",
                return_value={"kernels": [tunable]},
            ),
            mock.patch(
                "onnx_light.__main__._run_backend_test_timing", side_effect=run_timing
            ) as mocked,
            redirect_stdout(io.StringIO()),
            redirect_stderr(progress_output),
        ):
            main(
                [
                    "backend",
                    "--regex",
                    "gemm",
                    "--kernel",
                    "Gemm",
                    "--dtype",
                    "FLOAT",
                    "--impl",
                    "portable",
                    "--parameter",
                    "algorithm.tile=default,16",
                    "--parameter",
                    "parallel.minimum_tasks=default,4",
                    "--criterion",
                    "median-speedup",
                    "--json",
                ]
            )

        comparison = mocked.call_args.kwargs["tuning_comparison"]
        self.assertEqual(len(comparison["parameter_sets"]), 4)
        self.assertEqual(comparison["criterion"], "median-speedup")
        self.assertIn("[####################] 4/4", progress_output.getvalue())

    def test_saves_one_self_contained_model_per_test(self):
        from onnx_light.onnx import load

        with tempfile.TemporaryDirectory() as temporary:
            output = os.path.join(temporary, "models")
            report = _run_backend_test_timing(
                name_regex=r"^test_cc_not$", repeat=1, warmup=0, save_models=output
            )
            model_path = os.path.join(output, "test_cc_not.onnx")
            self.assertEqual(os.listdir(output), ["test_cc_not.onnx"])
            self.assertEqual(report["cases"][0]["model_path"], model_path)
            model = load(model_path)
            self.assertEqual(model.graph.name, "test_cc_not")

    def test_parameter_requires_tuning_schema_selectors(self):
        with self.assertRaisesRegex(
            SystemExit, "--parameter requires --kernel, --dtype, and --impl"
        ):
            main(
                [
                    "backend",
                    "--regex",
                    "not",
                    "--parameter",
                    "parallel.minimum_elements=default,16384",
                    "--criterion",
                    "sum",
                ]
            )

    def test_parameter_requires_criterion(self):
        with self.assertRaisesRegex(SystemExit, "--parameter requires --criterion"):
            main(
                [
                    "backend",
                    "--regex",
                    "not",
                    "--kernel",
                    "Not",
                    "--dtype",
                    "BOOL",
                    "--impl",
                    "portable",
                    "--parameter",
                    "parallel.minimum_elements=default,16384",
                ]
            )

    def test_parameter_speedup_is_unavailable_when_baseline_times_out(self):
        tunable = {
            "library": "onnx_light",
            "kernel": "Not",
            "implementation": "portable",
            "element_type": 9,
            "device": -1,
            "device_name": "CPU",
            "tuning_abi": 1,
            "calibratable": True,
            "defaults": {"parallel.minimum_elements": 32768},
            "active_values": {"parallel.minimum_elements": 32768},
            "parameter_names": ["parallel.minimum_elements"],
        }
        timeout_case = {
            "name": "test_cc_not",
            "kind": "node",
            "tag": "Not",
            "status": "timeout",
            "timed_out": True,
            "error": "exceeded 2 seconds",
            "data_sets": None,
            "input_shapes": None,
            "materialization_seconds": None,
            "setup_seconds": None,
            "warmup_seconds": None,
            "iteration_seconds": [],
            "run_seconds": None,
            "mean_seconds": None,
            "min_seconds": None,
            "max_seconds": None,
        }
        completed_case = {
            **timeout_case,
            "status": "completed",
            "timed_out": False,
            "error": None,
            "data_sets": 1,
            "iteration_seconds": [0.1],
            "run_seconds": 0.1,
            "mean_seconds": 0.1,
            "min_seconds": 0.1,
            "max_seconds": 0.1,
        }
        backend_case = mock.Mock()
        backend_case.name = "test_cc_not"
        backend_case.kind = "node"
        backend_case.tag = "Not"
        with (
            mock.patch(
                "onnx_light.onnx.backend.collect_test_cases_by_name", return_value=[backend_case]
            ),
            mock.patch(
                "onnx_light.__main__._measure_backend_test_cases_with_timeout",
                side_effect=[[timeout_case], [completed_case]],
            ),
            mock.patch("onnx_light.kernel_tuning.set_kernel_tuning_parameters"),
        ):
            report = _run_backend_test_timing(
                tuning_comparison={
                    "tunable": tunable,
                    "parameter_name": "parallel.minimum_elements",
                    "parameter_values": [32768, 16384],
                }
            )
        self.assertIsNone(report["tuning_comparison"]["values"][0]["speedup"])
        self.assertIsNone(report["tuning_comparison"]["values"][1]["speedup"])
        self.assertEqual(report["tuning_comparison"]["values"][1]["average"], 0.1)
        self.assertEqual(
            {
                "average",
                "sum",
                "median",
                "average_speedup",
                "median_speedup",
                "max_speedup",
                "max_latency",
            },
            {
                name
                for name in report["tuning_comparison"]["values"][1]
                if name
                in {
                    "average",
                    "sum",
                    "median",
                    "average_speedup",
                    "median_speedup",
                    "max_speedup",
                    "max_latency",
                }
            },
        )

    def test_rejects_unknown_output_extension(self):
        with (
            mock.patch(
                "onnx_light.__main__._run_backend_test_timing",
                return_value={
                    "mode": "test",
                    "repeat": 1,
                    "warmup": 0,
                    "collection_seconds": 0.0,
                    "cases": [],
                    "total_seconds": 0.0,
                },
            ),
            self.assertRaisesRegex(SystemExit, "expected .csv or .xlsx"),
        ):
            main(["backend", "--output", "report.json"])


if __name__ == "__main__":
    unittest.main()
